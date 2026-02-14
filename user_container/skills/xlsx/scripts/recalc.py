# /// script
# dependencies = ["openpyxl"]
# ///
"""
Excel Formula Recalculation Script

Recalculates all formulas in an Excel file using LibreOffice headless mode.
Returns JSON with error details if formula errors are found.

Usage:
    uv run recalc.py <excel_file> [timeout_seconds]

Example:
    uv run recalc.py financial_model.xlsx 60
"""

import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


MACRO_CONTENT = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''


def get_macro_dir() -> Path:
    """Get LibreOffice macro directory path."""
    return Path.home() / ".config/libreoffice/4/user/basic/Standard"


def setup_macro() -> bool:
    """
    Setup LibreOffice macro for recalculation.

    Creates the macro file if it doesn't exist or doesn't contain our macro.
    Returns True if setup successful, False otherwise.
    """
    macro_dir = get_macro_dir()
    macro_file = macro_dir / "Module1.xba"

    # Check if macro already exists
    if macro_file.exists():
        content = macro_file.read_text()
        if "RecalculateAndSave" in content:
            return True

    # Create directory if needed (may require running LibreOffice once)
    if not macro_dir.exists():
        try:
            subprocess.run(
                ["soffice", "--headless", "--terminate_after_init"],
                capture_output=True,
                timeout=15
            )
        except (subprocess.TimeoutExpired, FileNotFoundError):
            pass
        macro_dir.mkdir(parents=True, exist_ok=True)

    # Write macro file
    try:
        macro_file.write_text(MACRO_CONTENT)
        return True
    except Exception as e:
        print(f"Failed to write macro: {e}", file=sys.stderr)
        return False


def scan_for_errors(filepath: str) -> dict:
    """
    Scan Excel file for formula errors.

    Returns dict with error details including locations.
    """
    excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
    error_details = {err: [] for err in excel_errors}
    total_errors = 0

    try:
        wb = load_workbook(filepath, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        # Build error summary (only non-empty categories)
        error_summary = {}
        for err_type, locations in error_details.items():
            if locations:
                error_summary[err_type] = {
                    "count": len(locations),
                    "locations": locations[:50]  # Show up to 50 locations per type
                }

        return {
            "total_errors": total_errors,
            "error_summary": error_summary
        }

    except Exception as e:
        return {"scan_error": str(e)}


def count_formulas(filepath: str) -> int:
    """Count total formulas in workbook."""
    try:
        wb = load_workbook(filepath, data_only=False)
        count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        count += 1
        wb.close()
        return count
    except Exception:
        return -1


def recalc(filepath: str, timeout: int = 30) -> dict:
    """
    Recalculate all formulas in an Excel file using LibreOffice.

    Args:
        filepath: Path to Excel file (.xlsx, .xlsm)
        timeout: Maximum seconds to wait for recalculation

    Returns:
        dict with status, error counts, and locations
    """
    path = Path(filepath)

    if not path.exists():
        return {"error": f"File not found: {filepath}"}

    if not path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]:
        return {"error": f"Unsupported file type: {path.suffix}"}

    abs_path = str(path.absolute())

    # Setup macro if needed
    if not setup_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    # Build command with timeout
    cmd = [
        "timeout", str(timeout),
        "soffice", "--headless", "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path
    ]

    result = subprocess.run(cmd, capture_output=True, text=True)

    # Check for timeout (exit code 124)
    if result.returncode == 124:
        return {"error": f"Recalculation timed out after {timeout} seconds"}

    if result.returncode != 0:
        error_msg = result.stderr.strip() if result.stderr else "Unknown error"
        # Check for common issues
        if "Module1" in error_msg or "RecalculateAndSave" in error_msg:
            return {"error": "LibreOffice macro not configured. Try running: soffice --headless --terminate_after_init"}
        return {"error": f"LibreOffice error: {error_msg}"}

    # Scan for errors in recalculated file
    error_scan = scan_for_errors(abs_path)

    if "scan_error" in error_scan:
        return {"error": f"Failed to scan file: {error_scan['scan_error']}"}

    # Count formulas for context
    formula_count = count_formulas(abs_path)

    return {
        "status": "success" if error_scan["total_errors"] == 0 else "errors_found",
        "file": filepath,
        "total_errors": error_scan["total_errors"],
        "total_formulas": formula_count,
        "error_summary": error_scan["error_summary"]
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: uv run recalc.py <excel_file> [timeout_seconds]")
        print()
        print("Recalculates all formulas in an Excel file using LibreOffice.")
        print()
        print("Returns JSON with:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Number of formula errors (#VALUE!, #REF!, etc.)")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with cell locations")
        sys.exit(1)

    filepath = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filepath, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
