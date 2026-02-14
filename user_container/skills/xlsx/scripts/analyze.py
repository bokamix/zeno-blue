# /// script
# dependencies = ["openpyxl"]
# ///
"""
Excel File Inspector

Analyzes Excel file structure and returns detailed information about:
- Sheets (names, dimensions, row/column counts)
- Named ranges
- Formulas (count per sheet, complexity analysis)
- Data types distribution
- Merged cells
- Conditional formatting rules
- Data validations

Usage:
    uv run inspect.py <excel_file> [--formulas] [--full]

Options:
    --formulas  Include list of all formulas with locations
    --full      Include all details (formulas, validations, etc.)
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def inspect_sheet(ws, include_formulas: bool = False) -> dict:
    """Inspect a single worksheet."""
    info = {
        "name": ws.title,
        "dimensions": ws.dimensions or "Empty",
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": len(ws.merged_cells.ranges),
        "formulas": {"count": 0, "cells": []},
        "data_types": {"string": 0, "number": 0, "date": 0, "boolean": 0, "empty": 0, "formula": 0},
    }

    # Count conditional formatting rules
    info["conditional_formats"] = len(ws.conditional_formatting._cf_rules)

    # Count data validations
    info["data_validations"] = len(ws.data_validations.dataValidation) if ws.data_validations else 0

    # Analyze cells
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                info["data_types"]["empty"] += 1
            elif isinstance(cell.value, str):
                if cell.value.startswith("="):
                    info["data_types"]["formula"] += 1
                    info["formulas"]["count"] += 1
                    if include_formulas:
                        info["formulas"]["cells"].append({
                            "cell": cell.coordinate,
                            "formula": cell.value[:100]  # Truncate long formulas
                        })
                else:
                    info["data_types"]["string"] += 1
            elif isinstance(cell.value, bool):
                info["data_types"]["boolean"] += 1
            elif isinstance(cell.value, (int, float)):
                info["data_types"]["number"] += 1
            else:
                info["data_types"]["date"] += 1

    # Limit formula list
    if include_formulas and len(info["formulas"]["cells"]) > 100:
        info["formulas"]["cells"] = info["formulas"]["cells"][:100]
        info["formulas"]["truncated"] = True

    if not include_formulas:
        del info["formulas"]["cells"]

    return info


def inspect_workbook(filepath: str, include_formulas: bool = False, full: bool = False) -> dict:
    """Inspect an Excel workbook."""
    path = Path(filepath)

    if not path.exists():
        return {"error": f"File not found: {filepath}"}

    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        return {"error": f"Failed to open file: {e}"}

    result = {
        "file": filepath,
        "sheets_count": len(wb.sheetnames),
        "sheets": [],
        "named_ranges": [],
        "total_formulas": 0,
        "total_cells": 0,
    }

    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = inspect_sheet(ws, include_formulas=include_formulas or full)
        result["sheets"].append(sheet_info)
        result["total_formulas"] += sheet_info["formulas"]["count"]
        result["total_cells"] += sheet_info["max_row"] * sheet_info["max_column"]

    # Get named ranges
    for name in wb.defined_names:
        try:
            defn = wb.defined_names[name]
            result["named_ranges"].append({
                "name": name,
                "scope": defn.localSheetId if defn.localSheetId is not None else "global",
                "refers_to": defn.attr_text,
            })
        except Exception:
            result["named_ranges"].append({
                "name": name,
                "refers_to": "(error reading)",
            })

    wb.close()

    # Summary
    result["summary"] = {
        "total_sheets": result["sheets_count"],
        "total_formulas": result["total_formulas"],
        "total_named_ranges": len(result["named_ranges"]),
        "sheets_with_formulas": sum(1 for s in result["sheets"] if s["formulas"]["count"] > 0),
    }

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: uv run inspect.py <excel_file> [--formulas] [--full]")
        print()
        print("Analyzes Excel file structure and returns JSON with details about:")
        print("  - Sheets (dimensions, row/column counts)")
        print("  - Named ranges")
        print("  - Formula counts and locations")
        print("  - Data types distribution")
        print("  - Merged cells, conditional formatting, data validations")
        sys.exit(1)

    filepath = sys.argv[1]
    include_formulas = "--formulas" in sys.argv
    full = "--full" in sys.argv

    result = inspect_workbook(filepath, include_formulas=include_formulas, full=full)
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()
