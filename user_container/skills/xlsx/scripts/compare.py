# /// script
# dependencies = ["openpyxl", "pandas"]
# ///
"""
Excel File Comparator

Compare two Excel files and report differences:
- Sheet structure differences
- Cell value differences
- Formula differences
- Added/removed rows and columns

Usage:
    uv run compare.py <file1> <file2> [options]

Options:
    --sheet NAME    Compare specific sheet only
    --values        Compare calculated values (default: compare formulas)
    --summary       Show summary only, not individual differences
"""

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def compare_sheets(ws1, ws2, compare_values: bool = False) -> dict:
    """Compare two worksheets cell by cell."""
    differences = []

    max_row = max(ws1.max_row, ws2.max_row)
    max_col = max(ws1.max_column, ws2.max_column)

    cells_compared = 0
    cells_different = 0

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell1 = ws1.cell(row=row, column=col)
            cell2 = ws2.cell(row=row, column=col)

            val1 = cell1.value
            val2 = cell2.value

            cells_compared += 1

            # Skip if both empty
            if val1 is None and val2 is None:
                continue

            # Compare values
            if val1 != val2:
                cells_different += 1
                if len(differences) < 100:  # Limit to first 100 differences
                    differences.append({
                        "cell": cell1.coordinate,
                        "file1": str(val1)[:50] if val1 else "(empty)",
                        "file2": str(val2)[:50] if val2 else "(empty)",
                    })

    return {
        "cells_compared": cells_compared,
        "cells_different": cells_different,
        "match_percentage": round((1 - cells_different / max(cells_compared, 1)) * 100, 2),
        "differences": differences,
        "differences_truncated": cells_different > 100
    }


def compare_workbooks(file1: str, file2: str, sheet_name: str = None,
                      compare_values: bool = False, summary_only: bool = False) -> dict:
    """Compare two Excel workbooks."""
    path1 = Path(file1)
    path2 = Path(file2)

    if not path1.exists():
        return {"error": f"File not found: {file1}"}
    if not path2.exists():
        return {"error": f"File not found: {file2}"}

    try:
        wb1 = load_workbook(file1, data_only=compare_values)
        wb2 = load_workbook(file2, data_only=compare_values)
    except Exception as e:
        return {"error": f"Failed to open files: {e}"}

    result = {
        "file1": file1,
        "file2": file2,
        "compare_mode": "values" if compare_values else "formulas",
        "structure": {
            "file1_sheets": wb1.sheetnames,
            "file2_sheets": wb2.sheetnames,
            "sheets_only_in_file1": [s for s in wb1.sheetnames if s not in wb2.sheetnames],
            "sheets_only_in_file2": [s for s in wb2.sheetnames if s not in wb1.sheetnames],
            "common_sheets": [s for s in wb1.sheetnames if s in wb2.sheetnames],
        },
        "sheet_comparisons": []
    }

    # Determine which sheets to compare
    if sheet_name:
        if sheet_name not in wb1.sheetnames:
            return {"error": f"Sheet '{sheet_name}' not found in {file1}"}
        if sheet_name not in wb2.sheetnames:
            return {"error": f"Sheet '{sheet_name}' not found in {file2}"}
        sheets_to_compare = [sheet_name]
    else:
        sheets_to_compare = result["structure"]["common_sheets"]

    # Compare each sheet
    total_different = 0
    for sheet in sheets_to_compare:
        ws1 = wb1[sheet]
        ws2 = wb2[sheet]

        comparison = compare_sheets(ws1, ws2, compare_values)
        total_different += comparison["cells_different"]

        sheet_result = {
            "sheet": sheet,
            "cells_compared": comparison["cells_compared"],
            "cells_different": comparison["cells_different"],
            "match_percentage": comparison["match_percentage"],
        }

        if not summary_only:
            sheet_result["differences"] = comparison["differences"]
            if comparison["differences_truncated"]:
                sheet_result["differences_truncated"] = True

        result["sheet_comparisons"].append(sheet_result)

    wb1.close()
    wb2.close()

    # Overall summary
    total_compared = sum(s["cells_compared"] for s in result["sheet_comparisons"])
    result["summary"] = {
        "sheets_compared": len(sheets_to_compare),
        "total_cells_compared": total_compared,
        "total_cells_different": total_different,
        "overall_match_percentage": round((1 - total_different / max(total_compared, 1)) * 100, 2),
        "identical": total_different == 0 and not result["structure"]["sheets_only_in_file1"] and not result["structure"]["sheets_only_in_file2"]
    }

    return result


def main():
    if len(sys.argv) < 3:
        print("Usage: uv run compare.py <file1> <file2> [options]")
        print()
        print("Compare two Excel files and report differences.")
        print()
        print("Options:")
        print("  --sheet NAME    Compare specific sheet only")
        print("  --values        Compare calculated values (default: compare formulas)")
        print("  --summary       Show summary only, not individual differences")
        print()
        print("Output includes:")
        print("  - Sheet structure differences")
        print("  - Cell-by-cell comparison with locations")
        print("  - Match percentage per sheet and overall")
        sys.exit(1)

    file1 = sys.argv[1]
    file2 = sys.argv[2]

    sheet_name = None
    compare_values = "--values" in sys.argv
    summary_only = "--summary" in sys.argv

    # Parse --sheet option
    for i, arg in enumerate(sys.argv):
        if arg == "--sheet" and i + 1 < len(sys.argv):
            sheet_name = sys.argv[i + 1]

    result = compare_workbooks(file1, file2, sheet_name, compare_values, summary_only)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
